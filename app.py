import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title="Attendance Processor", page_icon="📊", layout="wide")

st.title("📊 Weekly Attendance Processor")

# Layout with two columns for uploads
col1, col2 = st.columns(2)

with col1:
    st.markdown("### <h1 style='font-size: 24px;'>1. Upload Master Sheet</h1>", unsafe_allow_html=True)
    master_file = st.file_uploader("Select the master Excel file", type=['xlsx'], label_visibility="collapsed")
    
    if master_file:
        st.subheader("Master Sheet Preview")
        preview_master = pd.read_excel(master_file).head(5)
        st.dataframe(preview_master, use_container_width=True)

with col2:
    st.markdown("### <h1 style='font-size: 24px;'>2. Upload Student Responses</h1>", unsafe_allow_html=True)
    response_file = st.file_uploader("Select the student responses file", type=['xlsx'], label_visibility="collapsed")
    
    if response_file:
        st.subheader("Responses Preview")
        preview_resp = pd.read_excel(response_file).head(5)
        st.dataframe(preview_resp, use_container_width=True)

if master_file and response_file:
    st.divider()
    if st.button("🚀 Process Attendance", type="primary"):
        try:
            # 1. Load responses and extract ID from Email
            df_responses = pd.read_excel(response_file)
            
            email_col = next((c for c in df_responses.columns if 'Email' in c), None)
            
            if email_col:
                # We create a mapping of normalized ID to original data for later display
                # Also clean the email column in the dataframe itself for easier filtering
                df_responses['temp_id'] = (df_responses[email_col]
                                           .astype(str)
                                           .str.split('@').str[0] # Access the first part of the split
                                           .str.replace('.0', '', regex=False)
                                           .str.strip()
                                           .str.lower())          # Use .str before .lower()
                response_ids = set(df_responses['temp_id'].unique())
            else:
                st.error("Could not find an 'Email' column in the Responses file.")
                st.stop()

            # 2. Process Master File using openpyxl
            master_file.seek(0)
            wb = openpyxl.load_workbook(master_file)
            ws = wb.active
            
            col_map = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value is not None}
            
            if "StudentNumber" not in col_map or "Status" not in col_map:
                st.error("Error: Master Sheet must have 'StudentNumber' and 'Status' columns.")
                st.stop()

            absentees_list = []
            present_count = 0
            master_ids_found = set() # Track which master IDs exist

            # 3. Iterate through Master Sheet rows
            for row in range(2, ws.max_row + 1):
                sid_cell = ws.cell(row=row, column=col_map["StudentNumber"])
                sid_val = sid_cell.value
                
                if sid_val is None: 
                    continue
                
                sid = str(sid_val).replace('.0', '').strip().lower()
                master_ids_found.add(sid) # Collect all valid IDs in master
                
                name_val = ws.cell(row=row, column=col_map.get("StudentName", col_map["StudentNumber"])).value

                if sid in response_ids:
                    ws.cell(row=row, column=col_map["Status"]).value = "Present"
                    present_count += 1
                else:
                    ws.cell(row=row, column=col_map["Status"]).value = "Absent"
                    absentees_list.append({
                        "Index": row - 1,
                        "StudentNumber": sid_val, 
                        "StudentName": name_val
                    })

            # --- LOGIC FOR UNRECOGNIZED STUDENTS ---
            # Find IDs that are in the Response file but NOT in the Master sheet
            unrecognized_ids = response_ids - master_ids_found
            df_unrecognized = df_responses[df_responses['temp_id'].isin(unrecognized_ids)].drop(columns=['temp_id'])

            # 4. Save modified workbook
            master_output_buffer = BytesIO()
            wb.save(master_output_buffer)
            master_output_data = master_output_buffer.getvalue()
            
            # Create Absentee List Excel
            absentee_df = pd.DataFrame(absentees_list)
            absentee_buffer = BytesIO()
            with pd.ExcelWriter(absentee_buffer, engine='openpyxl') as writer:
                absentee_df.to_excel(writer, index=False, sheet_name='Absentees')
            absentee_output_data = absentee_buffer.getvalue()

            # Results UI
            st.balloons()
            st.success(f"Processing Complete! {present_count} Present, {len(absentees_list)} Absent.")
            
            # Display Absentees
            if not absentee_df.empty:
                st.subheader("📋 Absentee List Summary")
                st.dataframe(absentee_df, use_container_width=True)

            # --- NEW DISPLAY SECTION ---
            # Display Unrecognized Students
            if not df_unrecognized.empty:
                st.warning(f"⚠️ Found {len(df_unrecognized)} student(s) who submitted responses but are NOT in the Master Sheet.")
                st.subheader("🔍 Unrecognized Students Details")
                st.dataframe(df_unrecognized, use_container_width=True)
            # ---------------------------

            st.divider()
            timestamp = datetime.now().strftime("%Y-%m-%d")
            
            dl_col1, dl_col2 = st.columns(2)
            with dl_col1:
                st.download_button(
                    label="📥 Download Attendance",
                    data=master_output_data,
                    file_name=f"Attendance_Status_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with dl_col2:
                st.download_button(
                    label="⚠️ Download Absentee List (Excel)",
                    data=absentee_output_data,
                    file_name=f"Absentees_Only_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        except Exception as e:
            st.error(f"An error occurred: {e}")
